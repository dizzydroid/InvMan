import sys
import os
import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QLabel, QPushButton, QLineEdit,
    QWidget, QFileDialog, QMessageBox, QScrollArea, QGridLayout, QDialog, QHBoxLayout, QComboBox, QDateEdit, QSizePolicy
)
from PyQt5.QtGui import QPixmap, QFont, QIcon
from PyQt5.QtCore import Qt, QDate
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import shutil

class InventoryApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("InvMan")
        self.setGeometry(100, 100, 800, 600)
        self.setWindowIcon(QIcon('icon.png'))  
        self.showMaximized()

        # Define the structure of the inventory
        self.inventory_file = 'inventory.xlsx'
        self.order_file = 'orders.xlsx'
        self.performance_file = 'performance.xlsx'  # New file for performance checks


        # Predefined categories
        self.categories = ["Cases", "Screen Protectors", "Chargers", "Headphones", "Speakers", "Cables", "Power Banks", "Mounts", "Stands", "Other"]

        # Create UI Elements
        self.initUI()

        # Load inventory data
        self.load_inventory()

    from PyQt5.QtWidgets import QSizePolicy

    def initUI(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QLabel {
                font-size: 14px;
            }
            QLineEdit {
                font-size: 14px;
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 5px;
                background-color: white;
            }
            QComboBox {
                font-size: 14px;
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 5px;
                background-color: white;
            }
            QPushButton {
                font-size: 14px;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
                text-transform: uppercase;
                font-family: Helvetica;
                letter-spacing: 0.8rem;
                cursor: hand;
            }
            QPushButton#addButton {
                background-color: #26425a;
                color: white;
                font-weight: bold;
                text-transform: uppercase;
                font-family: Helvetica;
                letter-spacing: 0.8rem;
            }
            QPushButton#selectImageButton {
                background-color: #20345c;     
                color: White  
            }
            QPushButton#addModelButton {
                background-color: #854d25;
                color: White
            }
            QPushButton#orderButton {
                background-color: #3c7a58;
                color: white;
            }
            QPushButton#editButton {
                background-color: #3c4c7a;
                color: white;
            }
            QPushButton#removeButton {
                background-color: #8B0000;
                color: white;
            }
            QPushButton#refundButton {
                background-color: #ad0328;
                color: white;
            }
            QPushButton#addColorButton {
                background-color: #20345c;
                color: White; 
            }
            QPushButton#viewDetailsButton {
                background-color: #3c7a6c;
                color: white;
            }
            QPushButton#trackPerformanceButton {
                background-color: #7d2201;
                color: white;
                font-weight: bold;
                text-transform: uppercase;
                font-family: Helvetica;
                letter-spacing: 0.8rem;
            }
            QPushButton#viewAllDetailsButton {
            background-color: #072e33;
            color: white;
            font-weight: bold;
            text-transform: uppercase;
            font-family: Helvetica;
            letter-spacing: 0.8rem;
            }
            QPushButton#bestWorstButton {
            background-color: #7d0101;
            color: white;
            font-weight: bold;
            text-transform: uppercase;
            font-family: Helvetica;
            letter-spacing: 0.8rem;
        }
        """)

        layout = QVBoxLayout()

        self.search_bar = QLineEdit(self)
        self.search_bar.setPlaceholderText("Search by name")
        self.search_bar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.search_bar.textChanged.connect(self.apply_filters)
        layout.addWidget(self.search_bar)

        self.search_phone_model_bar = QLineEdit(self)
        self.search_phone_model_bar.setPlaceholderText("Search by phone model")
        self.search_phone_model_bar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.search_phone_model_bar.textChanged.connect(self.apply_filters)
        layout.addWidget(self.search_phone_model_bar)

        self.category_filter = QComboBox(self)
        self.category_filter.addItem("(NONE)")
        self.category_filter.addItems(self.categories)
        self.category_filter.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.category_filter.currentIndexChanged.connect(self.apply_filters)
        layout.addWidget(self.category_filter)

        self.scroll_area = QScrollArea(self)
        self.scroll_area_widget_contents = QWidget()
        self.scroll_area.setWidget(self.scroll_area_widget_contents)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.inventory_layout = QGridLayout(self.scroll_area_widget_contents)
        layout.addWidget(self.scroll_area)

        button_layout = QHBoxLayout()
        add_button = QPushButton("Add Item", self)
        add_button.setObjectName("addButton")
        add_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        add_button.clicked.connect(self.open_add_item_window)
        button_layout.addWidget(add_button)
        
        view_all_details_button = QPushButton("View All Details", self)
        view_all_details_button.setObjectName("viewAllDetailsButton")
        view_all_details_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        view_all_details_button.clicked.connect(self.view_all_details)
        button_layout.addWidget(view_all_details_button)

        track_performance_button = QPushButton("Track Performance", self)
        track_performance_button.setObjectName("trackPerformanceButton")
        track_performance_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        track_performance_button.clicked.connect(self.open_performance_window)
        button_layout.addWidget(track_performance_button)

        best_worst_sellers_button = QPushButton("Best/Worst Sellers", self)
        best_worst_sellers_button.setObjectName("bestWorstButton")
        best_worst_sellers_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        best_worst_sellers_button.clicked.connect(self.view_best_worst_sellers)
        button_layout.addWidget(best_worst_sellers_button)

        layout.addLayout(button_layout)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        self.showMaximized()

    def ensure_images_directory(self):
        if not os.path.exists("images"):
            os.makedirs("images")

    def load_inventory(self):
        try:
            if os.path.exists(self.inventory_file):
                self.inventory_df = pd.read_excel(self.inventory_file)
                self.inventory_df['Data'] = self.inventory_df['Data'].apply(eval)  # Convert string representation of dict back to dict
            else:
                self.inventory_df = pd.DataFrame(columns=['Item Name', 'Category', 'Data', 'Image Path'])
            self.load_orders()
            self.update_inventory_view()
        except Exception as e:
            print(f"Error loading inventory: {e}")

    def save_inventory(self):
        try:
            self.inventory_df['Data'] = self.inventory_df['Data'].apply(str)  # Convert dict to string representation
            self.inventory_df.to_excel(self.inventory_file, index=False)
            self.inventory_df['Data'] = self.inventory_df['Data'].apply(eval)  # Convert back to dict for in-memory operations
            self.update_inventory_view()
        except Exception as e:
            print(f"Error saving inventory: {e}")

    def load_orders(self):
        try:
            if os.path.exists(self.order_file):
                self.orders_df = pd.read_excel(self.order_file)
                self.orders_df = self.orders_df[['Order Name','Product Name', 'Model', 'Color', 'Quantity', 'Date', 'Unit Price', 'Model Fee', 'Shipping Fee', 'Total Price', 'Net Profit', 'Status']]
            else:
                self.orders_df = pd.DataFrame(columns=['Order Name','Product Name', 'Model', 'Color', 'Quantity', 'Date', 'Unit Price', 'Model Fee', 'Shipping Fee', 'Total Price', 'Net Profit', 'Status'])
        except Exception as e:
            print(f"Error loading orders: {e}")
            
    def save_orders(self):
        try:
            self.orders_df.to_excel(self.order_file, index=False)
            self.format_orders_sheet()
        except Exception as e:
            print(f"Error saving orders: {e}")

    def update_inventory_view(self):
        try:
            for i in reversed(range(self.inventory_layout.count())):
                widget_to_remove = self.inventory_layout.itemAt(i).widget()
                self.inventory_layout.removeWidget(widget_to_remove)
                widget_to_remove.setParent(None)
            for index, row in self.inventory_df.iterrows():
                self.display_product(row, index)
        except Exception as e:
            print(f"Error updating inventory view: {e}")

    def display_product(self, row, index):
        try:
            item_name = row['Item Name']
            category = row['Category']
            image_path = row['Image Path']

            # Product container
            product_container = QWidget()
            product_layout = QVBoxLayout()

            # Product image
            if isinstance(image_path, str) and os.path.exists(image_path):
                pixmap = QPixmap(image_path)
                pixmap = pixmap.scaled(150, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation)  # Adjust size as needed
            else:
                pixmap = QPixmap(150, 150)  # Adjust size as needed

            image_label = QLabel(self)
            image_label.setPixmap(pixmap)
            image_label.mousePressEvent = lambda event, idx=index: self.show_product_options(idx)
            product_layout.addWidget(image_label)

            # Product details with larger font and bold name
            details = QLabel(f"<b>Name: {item_name}</b><br>Category: {category}", self)
            details.setStyleSheet("font-size: 18px; margin-top: 10px;")  # Adjust font size and margin as needed
            product_layout.addWidget(details)

            product_container.setLayout(product_layout)
            self.inventory_layout.addWidget(product_container, index // 3, index % 3)
        except Exception as e:
            print(f"Error displaying product: {e}")


    def show_product_options(self, index):
        try:
            product = self.inventory_df.iloc[index]
            self.options_window = QDialog(self)
            self.options_window.setWindowTitle(f"Options for {product['Item Name']}")
            self.options_window.setGeometry(200, 200, 400, 400)

            layout = QVBoxLayout()

            layout.addWidget(QLabel(f"Product: {product['Item Name']}", self))
            add_button = QPushButton("Add Stock", self)
            add_button.setObjectName("addButton")
            add_button.clicked.connect(lambda: self.add_to_count(index))
            layout.addWidget(add_button)

            edit_button = QPushButton("Edit", self)
            edit_button.setObjectName("editButton")
            edit_button.clicked.connect(lambda: self.edit_product_info(index))
            layout.addWidget(edit_button)
            
            view_details_button = QPushButton("View Details", self)
            view_details_button.setObjectName("viewDetailsButton")
            view_details_button.clicked.connect(lambda: self.view_details(index))
            layout.addWidget(view_details_button)

            order_button = QPushButton("Order", self)
            order_button.setObjectName("orderButton")
            order_button.clicked.connect(lambda: self.order_product(index))
            layout.addWidget(order_button)


            refund_button = QPushButton("Refund", self)
            refund_button.setObjectName("refundButton")
            refund_button.clicked.connect(lambda: self.refund_product(index))
            layout.addWidget(refund_button)

            remove_button = QPushButton("Remove", self)
            remove_button.setObjectName("removeButton")
            remove_button.clicked.connect(lambda: self.remove_product(index))
            layout.addWidget(remove_button)

            self.options_window.setLayout(layout)
            self.options_window.exec_()
        except Exception as e:
            print(f"Error showing product options: {e}")

    def view_details(self, index):
        try:
            product = self.inventory_df.iloc[index]
            details_window = QDialog(self)
            details_window.setWindowTitle(f"Details for {product['Item Name']}")
            details_window.setGeometry(300, 300, 500, 300)

            layout = QVBoxLayout()

            layout.addWidget(QLabel(f"<b>Name:</b> <span style='color:blue;'>{product['Item Name']}</span>", self))
            layout.addWidget(QLabel(f"<b>Category:</b> <span style='color:green;'>{product['Category']}</span>", self))

            self.model_combobox = QComboBox(details_window)
            self.model_combobox.addItems(product['Data'].keys())
            self.model_combobox.currentTextChanged.connect(lambda: self.update_model_details(product))
            layout.addWidget(self.model_combobox)

            self.model_details_layout = QVBoxLayout()
            layout.addLayout(self.model_details_layout)

            self.update_model_details(product)

            details_window.setLayout(layout)
            details_window.exec_()
        except Exception as e:
            print(f"Error viewing details: {e}")

    def update_model_details(self, product):
        for i in reversed(range(self.model_details_layout.count())):
            widget_to_remove = self.model_details_layout.itemAt(i).widget()
            self.model_details_layout.removeWidget(widget_to_remove)
            widget_to_remove.setParent(None)

        model = self.model_combobox.currentText()
        if model in product['Data']:
            model_data = product['Data'][model]
            price = model_data.get("Price", 0)
            fee = model_data.get("Fee", 0)
            total_units_sold = model_data.get("Units Sold", 0)
            colors = "\n".join([f"{color}: {stock}" for color, stock in model_data.get("Colors", {}).items()])

            details_text = (f"Model: {model}<br>"
                            f"Price: <span style='color:red;'>${price:.2f}</span><br>"
                            f"Fee: <span style='color:orange;'>${fee:.2f}</span><br>"
                            f"Total Units Sold: <span style='color:purple;'>{total_units_sold}</span><br>"
                            f"Colors and Stock:<br>{colors}")

            details_label = QLabel(details_text, self)
            details_label.setTextFormat(Qt.RichText)
            self.model_details_layout.addWidget(details_label)

            # Dropdown for 'units sold' for each color
            units_sold_combobox = QComboBox(self)
            units_sold_combobox.addItem("Select color to view units sold")
            for color, stock in model_data.get("Colors", {}).items():
                units_sold = model_data.get("Units Sold Colors", {}).get(color, 0)
                units_sold_combobox.addItem(f"{color}: {units_sold} units sold")
            
            self.model_details_layout.addWidget(units_sold_combobox)




    def refund_product(self, index):
        try:
            product = self.inventory_df.iloc[index]
            self.refund_window = QDialog(self)
            self.refund_window.setWindowTitle(f"Refund {product['Item Name']}")
            self.refund_window.setGeometry(200, 200, 400, 400)

            layout = QVBoxLayout()
            
            refund_layout = QGridLayout()

            refund_details_label = QLabel("Refund Details:", self.refund_window)
            
            self.refund_shipping_fee_entry = QLineEdit(self.refund_window)
            self.refund_shipping_fee_entry.setPlaceholderText("Refund Shipping Fee")
            layout.addWidget(self.refund_shipping_fee_entry)

            self.refund_quantity_entry = QLineEdit(self.refund_window)
            self.refund_quantity_entry.setPlaceholderText("Quantity to Refund")
            refund_layout.addWidget(refund_details_label, 0, 0)
            refund_layout.addWidget(self.refund_quantity_entry, 0, 1)
            
            self.refund_model_combobox = QComboBox(self.refund_window)
            self.refund_model_combobox.addItems(product['Data'].keys())
            self.refund_model_combobox.currentTextChanged.connect(lambda: self.update_refund_colors(product))
            refund_layout.addWidget(QLabel("Model", self.refund_window), 1, 0)
            refund_layout.addWidget(self.refund_model_combobox, 1, 1)

            self.refund_colors_combobox = QComboBox(self.refund_window)
            refund_layout.addWidget(QLabel("Color", self.refund_window), 2, 0)
            refund_layout.addWidget(self.refund_colors_combobox, 2, 1)
            self.update_refund_colors(product)

            layout.addLayout(refund_layout)
            refund_button = QPushButton("Refund", self.refund_window)
            refund_button.setObjectName("refundButton")
            refund_button.clicked.connect(lambda: self.process_refund(index))
            layout.addWidget(refund_button)

            self.refund_window.setLayout(layout)
            self.refund_window.exec_()
        except Exception as e:
            print(f"Error in refund_product: {e}")

    def process_refund(self, index):
        try:
            product = self.inventory_df.iloc[index]
            refund_quantity = self.refund_quantity_entry.text()
            selected_model = self.refund_model_combobox.currentText()
            selected_color = self.refund_colors_combobox.currentText()
            refund_shipping_fee = self.refund_shipping_fee_entry.text()

            if refund_quantity.isdigit() and int(refund_quantity) > 0:
                refund_quantity = int(refund_quantity)
                unit_price = product['Data'][selected_model]['Price']
                model_fee = product['Data'][selected_model].get('Fee', 0)

                if refund_shipping_fee.isdigit() and int(refund_shipping_fee) >= 0:
                    refund_shipping_fee = int(refund_shipping_fee)
                    total_price_without_shipping = unit_price * refund_quantity
                    total_price = refund_shipping_fee
                    net_profit = (unit_price - model_fee) * refund_quantity

                    # Update stock without checking if it is sufficient
                    product['Data'][selected_model]['Colors'][selected_color] += refund_quantity
                    self.save_inventory()

                    # Update the order sheet to mark it as refunded
                    order_name = f"Refund-{product['Item Name']}-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
                    new_order = pd.DataFrame([[order_name, product['Item Name'], selected_model, selected_color, -refund_quantity, datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p'), unit_price, model_fee, refund_shipping_fee, total_price, -net_profit, 'REFUNDED']], columns=['Order Name', 'Product Name', 'Model', 'Color', 'Quantity', 'Date', 'Unit Price', 'Model Fee', 'Shipping Fee', 'Total Price', 'Net Profit', 'Status'])
                    self.orders_df = pd.concat([self.orders_df, new_order], ignore_index=True)
                    self.save_orders()

                    QMessageBox.information(self, "Success", "Refund processed successfully.")
                    self.refund_window.close()
                else:
                    QMessageBox.warning(self, "Error", "Please enter a valid refund shipping fee.")
            else:
                QMessageBox.warning(self, "Error", "Please enter a valid refund quantity.")
        except Exception as e:
            print(f"Error processing refund: {e}")



    def update_refund_colors(self, product):
        model = self.refund_model_combobox.currentText()
        self.refund_colors_combobox.clear()
        if model in product['Data']:
            self.refund_colors_combobox.addItems(product['Data'][model]['Colors'].keys())

    
    def open_add_item_window(self):
        try:
            self.add_window = QDialog(self)
            self.add_window.setWindowTitle("Add New Item")
            self.add_window.setGeometry(200, 200, 400, 600)

            layout = QVBoxLayout()

            self.item_name_entry = QLineEdit(self.add_window)
            self.item_name_entry.setPlaceholderText("Item Name")
            layout.addWidget(self.item_name_entry)

            self.category_combobox = QComboBox(self.add_window)
            self.category_combobox.addItems(self.categories)
            layout.addWidget(self.category_combobox)

            self.image_path_entry = QLineEdit(self.add_window)
            self.image_path_entry.setPlaceholderText("Image Path")
            layout.addWidget(self.image_path_entry)
            select_image_button = QPushButton("Select Image", self.add_window)
            select_image_button.setObjectName("selectImageButton")
            select_image_button.clicked.connect(self.select_image)
            layout.addWidget(select_image_button)

            self.models_layout = QVBoxLayout()
            self.model_fields = []

            models_container = QWidget()
            models_container.setLayout(self.models_layout)
            scroll_area = QScrollArea(self.add_window)
            scroll_area.setWidgetResizable(True)
            scroll_area.setWidget(models_container)

            self.add_model_button = QPushButton("Add Product Model", self.add_window)
            self.add_model_button.setObjectName("addModelButton")
            self.add_model_button.clicked.connect(lambda: self.add_model_fields(self.models_layout, self.model_fields))
            layout.addWidget(scroll_area)
            layout.addWidget(self.add_model_button)

            add_item_button = QPushButton("Add Item", self.add_window)
            add_item_button.setObjectName("addButton")
            add_item_button.clicked.connect(self.add_item)
            layout.addWidget(add_item_button)

            self.add_window.setLayout(layout)
            self.add_window.exec_()
        except Exception as e:
            print(f"Error opening add item window: {e}")


    def add_model_fields(self, parent_layout, model_fields_list):
        try:
            model_layout = QVBoxLayout()

            model_name_entry = QLineEdit(self.add_window)
            model_name_entry.setPlaceholderText("Model Name")
            model_layout.addWidget(model_name_entry)

            model_price_entry = QLineEdit(self.add_window)
            model_price_entry.setPlaceholderText("Price")
            model_layout.addWidget(model_price_entry)

            model_fee_entry = QLineEdit(self.add_window)
            model_fee_entry.setPlaceholderText("Fee (optional)")
            model_layout.addWidget(model_fee_entry)

            colors_layout = QVBoxLayout()
            add_color_button = QPushButton("Add Color and Stock", self.add_window)
            add_color_button.setObjectName("addColorButton")
            add_color_button.clicked.connect(lambda _, cl=colors_layout: self.add_color_stock_fields(cl))
            model_layout.addLayout(colors_layout)
            model_layout.addWidget(add_color_button)

            parent_layout.addLayout(model_layout)
            model_fields_list.append((model_name_entry, model_price_entry, model_fee_entry, colors_layout))
        except Exception as e:
            print(f"Error adding model fields: {e}")



    def add_color_stock_fields(self, colors_layout):
        try:
            color_stock_layout = QHBoxLayout()

            color_entry = QLineEdit(self.add_window)
            color_entry.setPlaceholderText("Color")
            color_stock_layout.addWidget(color_entry)

            stock_entry = QLineEdit(self.add_window)
            stock_entry.setPlaceholderText("Stock")
            color_stock_layout.addWidget(stock_entry)

            colors_layout.addLayout(color_stock_layout)
        except Exception as e:
            print(f"Error adding color and stock fields: {e}")

    def add_item(self):
        try:
            item_name = self.item_name_entry.text()
            category = self.category_combobox.currentText()
            image_path = self.image_path_entry.text()
            data = {}

            valid = True

            for model_name_entry, model_price_entry, model_fee_entry, colors_layout in self.model_fields:
                model_name = model_name_entry.text()
                model_price = model_price_entry.text()
                model_fee = model_fee_entry.text()

                if model_name and model_price.replace('.', '', 1).isdigit():
                    model_price = float(model_price)
                    model_fee = float(model_fee) if model_fee else 0.0
                    colors = {}

                    for i in range(colors_layout.count()):
                        color_layout = colors_layout.itemAt(i).layout()
                        color_entry = color_layout.itemAt(0).widget()
                        stock_entry = color_layout.itemAt(1).widget()

                        color = color_entry.text()
                        stock = stock_entry.text()

                        if color and stock.isdigit():
                            stock_quantity = int(stock)
                            if stock_quantity >= 0:
                                colors[color] = stock_quantity
                            else:
                                valid = False
                                break
                        elif color or stock:
                            valid = False
                            break

                    if not colors:
                        valid = False

                    data[model_name] = {"Price": model_price, "Fee": model_fee, "Colors": colors}
                else:
                    valid = False
                    break

            if valid and item_name and category and data and os.path.exists(image_path):
                new_item = pd.DataFrame([[item_name, category, data, image_path]], columns=['Item Name', 'Category', 'Data', 'Image Path'])
                self.inventory_df = pd.concat([self.inventory_df, new_item], ignore_index=True)

                self.save_inventory()
                QMessageBox.information(self, "Success", f"Added {item_name} to inventory.")
                self.add_window.close()
            else:
                QMessageBox.warning(self, "Error", "Please enter valid item details and ensure at least one model with colors and stock.")
        except Exception as e:
            print(f"Error adding item: {e}")


    def add_to_count(self, index):
        try:
            product = self.inventory_df.iloc[index]
            self.add_stock_window = QDialog(self)
            self.add_stock_window.setWindowTitle("Add Stock")
            self.add_stock_window.setGeometry(200, 200, 400, 400)

            layout = QVBoxLayout()

            stock_layout = QGridLayout()

            self.add_stock_entries = {}

            row = 0
            for model, model_data in product['Data'].items():
                model_label = QLabel(f"Model: {model}", self.add_stock_window)
                stock_layout.addWidget(model_label, row, 0, 1, 2)
                row += 1

                for color in model_data['Colors']:
                    color_label = QLabel(color, self.add_stock_window)
                    stock_entry = QLineEdit(self.add_stock_window)
                    stock_entry.setPlaceholderText("Add Stock")
                    self.add_stock_entries[(model, color)] = stock_entry

                    stock_layout.addWidget(color_label, row, 0)
                    stock_layout.addWidget(stock_entry, row, 1)
                    row += 1

            layout.addLayout(stock_layout)

            confirm_button = QPushButton("Confirm", self.add_stock_window)
            confirm_button.setObjectName("confirmButton")
            confirm_button.setStyleSheet("""
                QPushButton#confirmButton {
                    background-color: #4CAF50;
                    color: white;
                    font-weight: bold;
                    text-transform: uppercase;
                    font-family: Helvetica;
                    letter-spacing: 0.8rem;
                    padding: 10px;
                }
            """)
            confirm_button.clicked.connect(lambda: self.confirm_add_stock(index))
            layout.addWidget(confirm_button, alignment=Qt.AlignCenter)

            self.add_stock_window.setLayout(layout)
            self.add_stock_window.exec_()
        except Exception as e:
            print(f"Error adding to count: {e}")



    def confirm_add_stock(self, index):
        try:
            product = self.inventory_df.iloc[index]
            for (model, color), stock_entry in self.add_stock_entries.items():
                stock = stock_entry.text()
                if stock.isdigit():
                    stock_quantity = int(stock)
                    if stock_quantity >= 0:
                        self.inventory_df.at[index, 'Data'][model]['Colors'][color] += stock_quantity
                    else:
                        QMessageBox.warning(self, "Error", "Please enter a non-negative stock quantity for the selected model and color.")
                        return
                elif stock:
                    QMessageBox.warning(self, "Error", "Please enter a valid stock quantity for the selected model and color.")
                    return

            self.save_inventory()
            self.add_stock_window.close()
            QMessageBox.information(self, "Success", "Stock updated successfully.")
        except Exception as e:
            print(f"Error confirming add stock: {e}")

    def edit_product_info(self, index):
        try:
            product = self.inventory_df.iloc[index]
            self.edit_window = QDialog(self)
            self.edit_window.setWindowTitle("Edit Product Info")
            self.edit_window.setGeometry(200, 200, 400, 600)

            layout = QVBoxLayout()

            self.edit_item_name_entry = QLineEdit(self.edit_window)
            self.edit_item_name_entry.setText(product['Item Name'])
            layout.addWidget(self.edit_item_name_entry)

            self.edit_category_combobox = QComboBox(self.edit_window)
            self.edit_category_combobox.addItems(self.categories)
            self.edit_category_combobox.setCurrentText(product['Category'])
            layout.addWidget(self.edit_category_combobox)

            self.edit_image_path_entry = QLineEdit(self.edit_window)
            self.edit_image_path_entry.setText(product['Image Path'])
            layout.addWidget(self.edit_image_path_entry)
            select_image_button = QPushButton("Select Image", self.edit_window)
            select_image_button.setObjectName("selectImageButton")
            select_image_button.clicked.connect(self.select_edit_image)
            layout.addWidget(select_image_button)

            self.edit_data_layout = QVBoxLayout()
            self.edit_model_fields = []

            for model, model_data in product['Data'].items():
                model_layout = QVBoxLayout()

                model_name_entry = QLineEdit(self.edit_window)
                model_name_entry.setText(model)
                model_layout.addWidget(model_name_entry)

                model_price_entry = QLineEdit(self.edit_window)
                model_price_entry.setText(str(model_data['Price']))
                model_layout.addWidget(model_price_entry)

                model_fee_entry = QLineEdit(self.edit_window)
                model_fee_entry.setText(str(model_data.get('Fee', 0)))
                model_layout.addWidget(model_fee_entry)

                colors_layout = QVBoxLayout()
                for color, stock in model_data['Colors'].items():
                    color_stock_layout = QHBoxLayout()

                    color_entry = QLineEdit(self.edit_window)
                    color_entry.setText(color)
                    color_stock_layout.addWidget(color_entry)

                    stock_entry = QLineEdit(self.edit_window)
                    stock_entry.setText(str(stock))
                    color_stock_layout.addWidget(stock_entry)

                    colors_layout.addLayout(color_stock_layout)

                add_color_button = QPushButton("Add Color and Stock", self.edit_window)
                add_color_button.clicked.connect(lambda _, cl=colors_layout: self.add_edit_color_stock_fields(cl))
                model_layout.addLayout(colors_layout)
                model_layout.addWidget(add_color_button)
                add_color_button.setObjectName("addColorButton")

                self.edit_data_layout.addLayout(model_layout)
                self.edit_model_fields.append((model_name_entry, model_price_entry, model_fee_entry, colors_layout))

            models_container = QWidget()
            models_container.setLayout(self.edit_data_layout)
            scroll_area = QScrollArea(self.edit_window)
            scroll_area.setWidgetResizable(True)
            scroll_area.setWidget(models_container)

            self.add_model_button = QPushButton("Add Phone Model", self.edit_window)
            self.add_model_button.setObjectName("addModelButton")
            self.add_model_button.clicked.connect(lambda: self.add_edit_model_fields(self.edit_data_layout, self.edit_model_fields))

            layout.addWidget(scroll_area)
            layout.addWidget(self.add_model_button)

            save_button = QPushButton("Save", self.edit_window)
            save_button.setObjectName("addButton")
            save_button.clicked.connect(lambda: self.save_product_info(index))
            layout.addWidget(save_button)

            self.edit_window.setLayout(layout)
            self.edit_window.exec_()
        except Exception as e:
            print(f"Error editing product info: {e}")





    def add_edit_model_fields(self, parent_layout, model_fields_list):
        try:
            model_layout = QVBoxLayout()

            model_name_entry = QLineEdit(self.edit_window)
            model_name_entry.setPlaceholderText("Model Name")
            model_layout.addWidget(model_name_entry)

            model_price_entry = QLineEdit(self.edit_window)
            model_price_entry.setPlaceholderText("Price")
            model_layout.addWidget(model_price_entry)

            model_fee_entry = QLineEdit(self.edit_window)
            model_fee_entry.setPlaceholderText("Fee (optional)")
            model_layout.addWidget(model_fee_entry)

            colors_layout = QVBoxLayout()
            add_color_button = QPushButton("Add Color and Stock", self.edit_window)
            add_color_button.setObjectName("addColorButton")
            add_color_button.clicked.connect(lambda _, cl=colors_layout: self.add_edit_color_stock_fields(cl))
            model_layout.addLayout(colors_layout)
            model_layout.addWidget(add_color_button)

            parent_layout.addLayout(model_layout)
            model_fields_list.append((model_name_entry, model_price_entry, model_fee_entry, colors_layout))
        except Exception as e:
            print(f"Error adding model fields: {e}")


    def add_edit_color_stock_fields(self, colors_layout):
        try:
            color_stock_layout = QHBoxLayout()

            color_entry = QLineEdit(self.edit_window)
            color_entry.setPlaceholderText("Color")
            color_stock_layout.addWidget(color_entry)

            stock_entry = QLineEdit(self.edit_window)
            stock_entry.setPlaceholderText("Stock")
            color_stock_layout.addWidget(stock_entry)

            colors_layout.addLayout(color_stock_layout)
        except Exception as e:
            print(f"Error adding color and stock fields: {e}")
        

    def save_product_info(self, index):
        try:
            item_name = self.edit_item_name_entry.text()
            category = self.edit_category_combobox.currentText()
            image_path = self.edit_image_path_entry.text()
            data = {}

            valid = True

            for model_name_entry, model_price_entry, model_fee_entry, colors_layout in self.edit_model_fields:
                model_name = model_name_entry.text()
                model_price = model_price_entry.text()
                model_fee = model_fee_entry.text()

                if model_name and model_price.replace('.', '', 1).isdigit():
                    model_price = float(model_price)
                    model_fee = float(model_fee) if model_fee else 0.0
                    colors = {}

                    for i in range(colors_layout.count()):
                        color_layout = colors_layout.itemAt(i).layout()
                        color_entry = color_layout.itemAt(0).widget()
                        stock_entry = color_layout.itemAt(1).widget()

                        color = color_entry.text()
                        stock = stock_entry.text()

                        if color and stock.isdigit():
                            stock_quantity = int(stock)
                            if stock_quantity >= 0:
                                colors[color] = stock_quantity
                            else:
                                valid = False
                                break
                        elif color or stock:
                            valid = False
                            break

                    if not colors:
                        valid = False

                    data[model_name] = {"Price": model_price, "Fee": model_fee, "Colors": colors}
                else:
                    valid = False
                    break

            if valid and item_name and category and data and os.path.exists(image_path):
                self.inventory_df.at[index, 'Item Name'] = item_name
                self.inventory_df.at[index, 'Category'] = category
                self.inventory_df.at[index, 'Data'] = data
                self.inventory_df.at[index, 'Image Path'] = image_path

                self.save_inventory()
                QMessageBox.information(self, "Success", "Product information updated.")
                self.edit_window.close()
            else:
                QMessageBox.warning(self, "Error", "Please enter valid item details and ensure at least one model with colors and stock.")
        except Exception as e:
            print(f"Error saving product info: {e}")



    def select_edit_image(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Image Files (*.png *.jpg *.jpeg)")
            if file_path:
                self.ensure_images_directory()  # Ensure the images directory exists
                image_filename = os.path.basename(file_path)
                target_path = os.path.join("images", image_filename)
                shutil.copy(file_path, target_path)  # Copy the image to the images folder
                self.edit_image_path_entry.setText(target_path)  # Set the relative path
        except Exception as e:
            print(f"Error selecting edit image: {e}")


    def order_product(self, index):
        try:
            product = self.inventory_df.iloc[index]
            self.order_window = QDialog(self)
            self.order_window.setWindowTitle(f"Order {product['Item Name']}")
            self.order_window.setGeometry(200, 200, 400, 400)

            layout = QVBoxLayout()

            order_layout = QGridLayout()

            shipping_fee_label = QLabel("Shipping Fee (optional):", self.order_window)
            self.shipping_fee_entry = QLineEdit(self.order_window)
            self.shipping_fee_entry.setPlaceholderText("Shipping Fee")
            order_layout.addWidget(shipping_fee_label, 0, 0)
            order_layout.addWidget(self.shipping_fee_entry, 0, 1)

            quantity_label = QLabel("Quantity to Order", self.order_window)
            self.order_quantity_entry = QLineEdit(self.order_window)
            self.order_quantity_entry.setPlaceholderText("Quantity to Order")
            order_layout.addWidget(quantity_label, 1, 0)
            order_layout.addWidget(self.order_quantity_entry, 1, 1)

            order_name_label = QLabel("Order Name (optional)", self.order_window)
            self.order_name_entry = QLineEdit(self.order_window)
            self.order_name_entry.setPlaceholderText("Order Name (optional)")
            order_layout.addWidget(order_name_label, 2, 0)
            order_layout.addWidget(self.order_name_entry, 2, 1)

            self.order_model_combobox = QComboBox(self.order_window)
            self.order_model_combobox.addItems(product['Data'].keys())
            self.order_model_combobox.currentTextChanged.connect(lambda: self.update_order_colors(product))
            order_layout.addWidget(QLabel("Model", self.order_window), 3, 0)
            order_layout.addWidget(self.order_model_combobox, 3, 1)

            self.order_colors_combobox = QComboBox(self.order_window)
            order_layout.addWidget(QLabel("Color", self.order_window), 4, 0)
            order_layout.addWidget(self.order_colors_combobox, 4, 1)
            self.update_order_colors(product)

            layout.addLayout(order_layout)
            order_button = QPushButton("Order", self.order_window)
            order_button.setObjectName("orderButton")
            order_button.clicked.connect(lambda: self.generate_receipt(index))
            layout.addWidget(order_button)

            self.order_window.setLayout(layout)
            self.order_window.exec_()
        except Exception as e:
            print(f"Error in order_product: {e}")

    def update_order_colors(self, product):
        model = self.order_model_combobox.currentText()
        self.order_colors_combobox.clear()
        if model in product['Data']:
            self.order_colors_combobox.addItems(product['Data'][model]['Colors'].keys())

    def generate_receipt(self, index):
        try:
            product = self.inventory_df.iloc[index]
            shipping_fee = self.shipping_fee_entry.text()
            order_quantity = self.order_quantity_entry.text()
            order_name = self.order_name_entry.text() or "No Name"
            selected_model = self.order_model_combobox.currentText()
            selected_color = self.order_colors_combobox.currentText()
            order_date = datetime.datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")

            if order_quantity.isdigit() and (shipping_fee.replace('.', '', 1).isdigit() or shipping_fee == ""):
                order_quantity = int(order_quantity)
                unit_price = product['Data'][selected_model]['Price']
                model_fee = product['Data'][selected_model].get('Fee', 0)
                total_price_without_shipping = unit_price * order_quantity

                if shipping_fee:
                    shipping_fee = float(shipping_fee)
                else:
                    shipping_fee = 0.0

                total_price = total_price_without_shipping + shipping_fee
                net_profit = (unit_price - model_fee) * order_quantity

                if product['Data'][selected_model]['Colors'][selected_color] >= order_quantity:
                    self.inventory_df.at[index, 'Data'][selected_model]['Colors'][selected_color] -= order_quantity

                    # Update 'Units Sold' for the model and color
                    if 'Units Sold' in self.inventory_df.at[index, 'Data'][selected_model]:
                        self.inventory_df.at[index, 'Data'][selected_model]['Units Sold'] += order_quantity
                    else:
                        self.inventory_df.at[index, 'Data'][selected_model]['Units Sold'] = order_quantity

                    if 'Units Sold Colors' not in self.inventory_df.at[index, 'Data'][selected_model]:
                        self.inventory_df.at[index, 'Data'][selected_model]['Units Sold Colors'] = {}

                    if selected_color in self.inventory_df.at[index, 'Data'][selected_model]['Units Sold Colors']:
                        self.inventory_df.at[index, 'Data'][selected_model]['Units Sold Colors'][selected_color] += order_quantity
                    else:
                        self.inventory_df.at[index, 'Data'][selected_model]['Units Sold Colors'][selected_color] = order_quantity

                    self.save_inventory()

                    new_order = pd.DataFrame([[order_name, product['Item Name'], selected_model, selected_color, order_quantity, order_date, unit_price, model_fee, shipping_fee, total_price, net_profit, 'ORDERED']], columns=['Order Name', 'Product Name', 'Model', 'Color', 'Quantity', 'Date', 'Unit Price', 'Model Fee', 'Shipping Fee', 'Total Price', 'Net Profit', 'Status'])
                    self.orders_df = pd.concat([self.orders_df, new_order], ignore_index=True)
                    self.save_orders()

                    receipt = f"Order Name: {order_name}\nProduct Name: {product['Item Name']}\nModel: {selected_model}\nColor: {selected_color}\nQuantity: {order_quantity}\nDate: {order_date}\nUnit Price: ${unit_price:.2f}\nModel Fee: ${model_fee:.2f}\nShipping Fee: ${shipping_fee:.2f}\nTotal Price: ${total_price:.2f}\nNet Profit: ${net_profit:.2f}\nStatus: ORDERED"
                    QMessageBox.information(self, "Receipt", receipt)
                    self.order_window.close()
                else:
                    QMessageBox.warning(self, "Error", "Insufficient stock for the order.")
            else:
                QMessageBox.warning(self, "Error", "Please enter valid order details.")
        except Exception as e:
            print(f"Error in generate_receipt: {e}")



    def remove_product(self, index):
        try:
            confirm = QMessageBox.question(self, "Confirm Delete", "Are you sure you want to remove this product?", QMessageBox.Yes | QMessageBox.No)
            if confirm == QMessageBox.Yes:
                self.inventory_df.drop(index, inplace=True)
                self.save_inventory()
                self.options_window.close()
                QMessageBox.information(self, "Success", "Product removed from inventory.")
        except Exception as e:
            print(f"Error removing product: {e}")

    def apply_filters(self):
        try:
            filtered_df = self.inventory_df.copy()

            query = self.search_bar.text().lower()
            phone_model_query = self.search_phone_model_bar.text().lower()
            category = self.category_filter.currentText()

            if query and query != "search by name":
                filtered_df = filtered_df[filtered_df['Item Name'].str.lower().str.contains(query)]

            if phone_model_query and phone_model_query != "search by phone model":
                filtered_df = filtered_df[filtered_df['Data'].apply(lambda x: any(phone_model_query in model.lower() for model in x.keys()))]

            if category and category != "(NONE)":
                filtered_df = filtered_df[filtered_df['Category'] == category]

            self.display_filtered_inventory(filtered_df)
        except Exception as e:
            print(f"Error applying filters: {e}")

    def display_filtered_inventory(self, df):
        try:
            for i in reversed(range(self.inventory_layout.count())):
                widget_to_remove = self.inventory_layout.itemAt(i).widget()
                self.inventory_layout.removeWidget(widget_to_remove)
                widget_to_remove.setParent(None)

            for index, row in df.iterrows():
                self.display_product(row, index)
        except Exception as e:
            print(f"Error displaying filtered inventory: {e}")

    def open_performance_window(self):
        try:
            self.performance_window = QDialog(self)
            self.performance_window.setWindowTitle("Track Performance")
            self.performance_window.setGeometry(300, 300, 400, 200)

            layout = QVBoxLayout()

            self.start_date_edit = QDateEdit(self.performance_window)
            self.start_date_edit.setCalendarPopup(True)
            self.start_date_edit.setDisplayFormat("dd/MM/yyyy")
            self.start_date_edit.setDate(QDate.currentDate().addMonths(-1))
            layout.addWidget(QLabel("Start Date:"))
            layout.addWidget(self.start_date_edit)

            self.end_date_edit = QDateEdit(self.performance_window)
            self.end_date_edit.setCalendarPopup(True)
            self.end_date_edit.setDisplayFormat("dd/MM/yyyy")
            self.end_date_edit.setDate(QDate.currentDate())
            layout.addWidget(QLabel("End Date:"))
            layout.addWidget(self.end_date_edit)

            track_button = QPushButton("Track Performance", self.performance_window)
            track_button.setObjectName("trackPerformanceButton")
            track_button.clicked.connect(self.track_performance)
            layout.addWidget(track_button)

            self.performance_window.setLayout(layout)
            self.performance_window.exec_()
        except Exception as e:
            print(f"Error opening performance window: {e}")

    def view_all_details(self):
        try:
            details_window = QDialog(self)
            details_window.setWindowTitle("All Product Details")
            details_window.setGeometry(300, 300, 800, 600)

            layout = QVBoxLayout()

            scroll_area = QScrollArea(details_window)
            scroll_area.setWidgetResizable(True)
            scroll_content = QWidget()
            scroll_layout = QVBoxLayout(scroll_content)

            all_details_text = ""

            for index, product in self.inventory_df.iterrows():
                all_details_text += f"<b>Product:</b> {product['Item Name']}<br>"
                all_details_text += f"<b>Category:</b> {product['Category']}<br>"
                for model, model_data in product['Data'].items():
                    all_details_text += f"<b>Model:</b> {model}<br>"
                    all_details_text += f"<b>Price:</b> ${model_data['Price']}<br>"
                    all_details_text += f"<b>Fee:</b> ${model_data.get('Fee', 0)}<br>"
                    all_details_text += f"<b>Units Sold:</b> {model_data.get('Units Sold', 0)}<br>"
                    all_details_text += "<b>Colors and Stock:</b><br>"
                    for color, stock in model_data['Colors'].items():
                        all_details_text += f"{color}: {stock}<br>"
                    all_details_text += "<br>"

            details_label = QLabel(all_details_text)
            details_label.setTextFormat(Qt.RichText)
            scroll_layout.addWidget(details_label)

            scroll_area.setWidget(scroll_content)
            layout.addWidget(scroll_area)

            details_window.setLayout(layout)
            details_window.exec_()
        except Exception as e:
            print(f"Error viewing all details: {e}")


    def view_best_worst_sellers(self):
        try:
            best_worst_window = QDialog(self)
            best_worst_window.setWindowTitle("Best and Worst Sellers")
            best_worst_window.setGeometry(300, 300, 800, 600)

            layout = QVBoxLayout()

            scroll_area = QScrollArea(best_worst_window)
            scroll_area.setWidgetResizable(True)
            scroll_content = QWidget()
            scroll_layout = QVBoxLayout(scroll_content)

            all_sales = []
            for index, row in self.inventory_df.iterrows():
                product_name = row['Item Name']
                for model, model_data in row['Data'].items():
                    for color, units_sold in model_data.get('Units Sold Colors', {}).items():
                        all_sales.append({
                            'Product Name': product_name,
                            'Model': model,
                            'Color': color,
                            'Units Sold': units_sold
                        })

            sales_df = pd.DataFrame(all_sales)
            best_sellers = sales_df.nlargest(3, 'Units Sold')
            worst_sellers = sales_df.nsmallest(3, 'Units Sold')

            best_sellers_text = "<b>Best Sellers:</b><br>"
            for index, row in best_sellers.iterrows():
                best_sellers_text += f"{row['Product Name']} - {row['Model']} ({row['Color']})<br>Units Sold: {row['Units Sold']}<br><br>"

            worst_sellers_text = "<b>Worst Sellers:</b><br>"
            for index, row in worst_sellers.iterrows():
                worst_sellers_text += f"{row['Product Name']} - {row['Model']} ({row['Color']})<br>Units Sold: {row['Units Sold']}<br><br>"

            details_label = QLabel(best_sellers_text + "<br>" + worst_sellers_text)
            details_label.setTextFormat(Qt.RichText)
            scroll_layout.addWidget(details_label)

            scroll_area.setWidget(scroll_content)
            layout.addWidget(scroll_area)

            best_worst_window.setLayout(layout)
            best_worst_window.exec_()
        except Exception as e:
            print(f"Error viewing best/worst sellers: {e}")



    def view_best_worst_colors(self):
        try:
            best_worst_window = QDialog(self)
            best_worst_window.setWindowTitle("Best and Worst Sellers by Color")
            best_worst_window.setGeometry(300, 300, 800, 600)

            layout = QVBoxLayout()

            scroll_area = QScrollArea(best_worst_window)
            scroll_area.setWidgetResizable(True)
            scroll_content = QWidget()
            scroll_layout = QVBoxLayout(scroll_content)

            all_colors = []
            for index, row in self.inventory_df.iterrows():
                product_name = row['Item Name']
                for model, model_data in row['Data'].items():
                    for color, units_sold in model_data.get('Units Sold Colors', {}).items():
                        all_colors.append({
                            'Product Name': product_name,
                            'Model': model,
                            'Color': color,
                            'Units Sold': units_sold
                        })

            colors_df = pd.DataFrame(all_colors)
            best_sellers = colors_df.nlargest(3, 'Units Sold')
            worst_sellers = colors_df.nsmallest(3, 'Units Sold')

            best_sellers_text = "<b>Best Sellers by Color:</b><br>"
            for index, row in best_sellers.iterrows():
                best_sellers_text += f"Product: {row['Product Name']}<br>Model: {row['Model']}<br>Color: {row['Color']}<br>Units Sold: {row['Units Sold']}<br><br>"

            worst_sellers_text = "<b>Worst Sellers by Color:</b><br>"
            for index, row in worst_sellers.iterrows():
                worst_sellers_text += f"Product: {row['Product Name']}<br>Model: {row['Model']}<br>Color: {row['Color']}<br>Units Sold: {row['Units Sold']}<br><br>"

            details_label = QLabel(best_sellers_text + "<br>" + worst_sellers_text)
            details_label.setTextFormat(Qt.RichText)
            scroll_layout.addWidget(details_label)

            scroll_area.setWidget(scroll_content)
            layout.addWidget(scroll_area)

            best_worst_window.setLayout(layout)
            best_worst_window.exec_()
        except Exception as e:
            print(f"Error viewing best/worst sellers by color: {e}")


    def format_orders_sheet(self):
        try:
            # Load the workbook and select the sheet
            workbook = openpyxl.load_workbook(self.order_file)
            sheet = workbook.active

            # Auto-fit column widths
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
            # Highlight "REFUND" rows
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for row in sheet.iter_rows():
                if row[-1].value == "REFUNDED":
                    for cell in row:
                        cell.fill = red_fill

            # Save the workbook with the applied formatting
            workbook.save(self.order_file)
        except Exception as e:
            print(f"Error formatting orders sheet: {e}")
    
    def track_performance(self):
        try:
            if not hasattr(self, 'orders_df') or self.orders_df.empty:
                QMessageBox.warning(self, "Error", "No orders to track performance.")
                return

            # Convert QDate to datetime
            start_date = self.start_date_edit.date().toPyDate()
            end_date = self.end_date_edit.date().toPyDate()

            # Convert to datetime with time for comparison
            start_date = pd.to_datetime(start_date)
            end_date = pd.to_datetime(end_date) + pd.Timedelta(days=1)

            print(f"Tracking performance from {start_date} to {end_date}")

            # Ensure Date column is in datetime format
            self.orders_df['Date'] = pd.to_datetime(self.orders_df['Date'], format='%Y-%m-%d %I:%M:%S %p')

            new_orders = self.orders_df[
                (self.orders_df['Status'] == 'ORDERED') &
                (self.orders_df['Date'].between(start_date, end_date))
            ]
            new_refunds = self.orders_df[
                (self.orders_df['Status'] == 'REFUNDED') &
                (self.orders_df['Date'].between(start_date, end_date))
            ]

            print("New Orders:")
            print(new_orders)
            print("New Refunds:")
            print(new_refunds)

            total_revenue = new_orders['Total Price'].sum()
            total_refunds = new_refunds['Total Price'].sum()
            net_profit = total_revenue + total_refunds  # Refunds are stored as negative values

            performance_entry = pd.DataFrame([{
                'Start Date': start_date.strftime('%d/%m/%Y'),
                'End Date': end_date.strftime('%d/%m/%Y'),
                'Net Profit': net_profit,
                'Tracked On': datetime.datetime.now().strftime('%d/%m/%Y %I:%M:%S %p')
            }])

            self.save_performance(performance_entry)

            QMessageBox.information(self, "Performance Tracked", f"Performance tracked from {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}.\nNet Profit: ${net_profit:.2f}")
            self.performance_window.close()
        except Exception as e:
            print(f"Error tracking performance: {e}")

    def save_performance(self, performance_entry):
        try:
            performance_file = self.performance_file
            if os.path.exists(performance_file):
                performance_df = pd.read_excel(performance_file)
                performance_df = pd.concat([performance_df, performance_entry], ignore_index=True)
            else:
                performance_df = performance_entry

            performance_df.to_excel(performance_file, index=False)

            self.format_performance_sheet(performance_file)
        except Exception as e:
            print(f"Error saving performance: {e}")

    def format_performance_sheet(self, performance_file):
        try:
            workbook = openpyxl.load_workbook(performance_file)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            workbook.save(performance_file)
        except Exception as e:
            print(f"Error formatting performance sheet: {e}")

    def select_image(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Image Files (*.png *.jpg *.jpeg)")
            if file_path:
                self.ensure_images_directory()  # Ensure the images directory exists
                image_filename = os.path.basename(file_path)
                target_path = os.path.join("images", image_filename)
                shutil.copy(file_path, target_path)  # Copy the image to the images folder
                self.image_path_entry.setText(target_path)  # Set the relative path
        except Exception as e:
            print(f"Error selecting image: {e}")


if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        window = InventoryApp()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"Error in main: {e}")
